from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse
import io

from .models import Personne, CATEGORIE_CHOICES
from .forms import PersonneForm, FiltreForm


def get_queryset_filtre(request):
    qs = Personne.objects.all()
    form = FiltreForm(request.GET)
    if form.is_valid():
        if form.cleaned_data.get('categorie'):
            qs = qs.filter(categorie=form.cleaned_data['categorie'])
        if form.cleaned_data.get('wilaya'):
            qs = qs.filter(wilaya__icontains=form.cleaned_data['wilaya'])
        if form.cleaned_data.get('recherche'):
            q = form.cleaned_data['recherche']
            qs = qs.filter(Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(nin__icontains=q))
        if form.cleaned_data.get('actif') != '':
            actif_val = form.cleaned_data['actif'] == '1'
            qs = qs.filter(actif=actif_val)
    return qs, form


@login_required
def accueil(request):
    stats = {
        'total': Personne.objects.count(),
        'mandoubs': Personne.objects.filter(categorie='mandoub').count(),
        'ouvriers': Personne.objects.filter(categorie='ouvrier').count(),
        'charges': Personne.objects.filter(categorie='charge_application').count(),
        'SG':Personne.objects.filter(categorie='SG').count(),
        'responsables': Personne.objects.filter(categorie='responsable').count(),
        'a_disposition': Personne.objects.filter(categorie='a_disposition').count(),
        'actifs': Personne.objects.filter(actif=True).count(),
    }
    derniers = Personne.objects.order_by('-cree_le')[:5]
    return render(request, 'personnel/accueil.html', {'stats': stats, 'derniers': derniers})


@login_required
def liste_personnes(request):
    qs, form = get_queryset_filtre(request)
    return render(request, 'personnel/liste.html', {
        'personnes': qs,
        'form': form,
        'total': qs.count(),
        'categories': CATEGORIE_CHOICES,
    })


@login_required
def ajouter_personne(request):
    if request.method == 'POST':
        form = PersonneForm(request.POST, request.FILES)
        if form.is_valid():
            personne = form.save()
            messages.success(request, f'تم إضافة {personne.get_full_name()} بنجاح ✓')
            return redirect('liste_personnes')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء أدناه.')
    else:
        cat = request.GET.get('categorie', '')
        form = PersonneForm(initial={'categorie': cat} if cat else {})
    return render(request, 'personnel/formulaire.html', {'form': form, 'titre': 'إضافة شخص جديد', 'mode': 'add'})


@login_required
def modifier_personne(request, pk):
    personne = get_object_or_404(Personne, pk=pk)
    if request.method == 'POST':
        form = PersonneForm(request.POST, request.FILES, instance=personne)
        if form.is_valid():
            form.save()
            messages.success(request, f'تم تعديل بيانات {personne.get_full_name()} بنجاح ✓')
            return redirect('detail_personne', pk=pk)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء أدناه.')
    else:
        form = PersonneForm(instance=personne)
    return render(request, 'personnel/formulaire.html', {
        'form': form,
        'titre': f'تعديل بيانات: {personne.get_full_name()}',
        'personne': personne,
        'mode': 'edit',
    })


@login_required
def detail_personne(request, pk):
    personne = get_object_or_404(Personne, pk=pk)
    return render(request, 'personnel/detail.html', {'personne': personne})


@user_passes_test(lambda u: u.is_staff or u.is_superuser)
@login_required
def supprimer_personne(request, pk):
    personne = get_object_or_404(Personne, pk=pk)
    if request.method == 'POST':
        nom = personne.get_full_name()
        personne.delete()
        messages.success(request, f'تم حذف {nom} بنجاح.')
        return redirect('liste_personnes')
    return render(request, 'personnel/confirmer_suppression.html', {'personne': personne})

@user_passes_test(lambda u: u.is_staff or u.is_superuser)
@login_required
def exporter_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة. قم بتشغيل: pip install openpyxl')
        return redirect('liste_personnes')

    qs, _ = get_queryset_filtre(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قائمة الأشخاص"
    ws.sheet_view.rightToLeft = True

    # Style header
    header_fill = PatternFill("solid", fgColor="1a5276")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'الرقم', 'الفئة', 'اللقب', 'الاسم', 'الجنس',
        'تاريخ الميلاد', 'مكان الميلاد', 'NIN',
        'الهاتف', 'البريد الإلكتروني',
        'العنوان', 'البلدية', 'الولاية',
        'RIB', 'البنك',
        'المنصب', 'المهمة',
        'تاريخ البداية', 'تاريخ النهاية',
        'صلاحية النظام', 'نشط', 'تاريخ الإنشاء',
    ]

    col_widths = [6, 20, 18, 18, 8, 14, 20, 22, 14, 26, 30, 16, 16, 26, 18, 22, 30, 14, 14, 16, 8, 18]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30

    # Alternating row colors
    fill_pair = [PatternFill("solid", fgColor="EBF5FB"), PatternFill("solid", fgColor="FFFFFF")]
    data_align = Alignment(horizontal="right", vertical="center")

    for row_num, p in enumerate(qs, 2):
        fill = fill_pair[row_num % 2]
        row_data = [
            row_num - 1,
            p.get_categorie_display(),
            p.nom, p.prenom,
            p.get_sexe_display(),
            p.date_naissance.strftime('%d/%m/%Y') if p.date_naissance else '',
            p.lieu_naissance, p.nin,
            p.telephone, p.email or '',
            p.adresse, p.commune,
            dict(p._meta.get_field('wilaya').choices).get(p.wilaya, p.wilaya),
            p.rib or '', p.banque or '',
            p.poste, p.mission or '',
            p.date_debut.strftime('%d/%m/%Y') if p.date_debut else '',
            p.date_fin.strftime('%d/%m/%Y') if p.date_fin else '',
            p.get_role_systeme_display(),
            'نعم' if p.actif else 'لا',
            p.cree_le.strftime('%d/%m/%Y') if p.cree_le else '',
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = fill
            cell.alignment = data_align
            cell.border = border

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto filter
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="قائمة_الافراد.xlsx"'
    return response
