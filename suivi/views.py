from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import models
import json, io

from .models import SuiviBaldiya, COMMUNE_CHOICES, STATUT_CHOICES, Tache, SuiviTache
from .forms import SuiviBaldiyaForm


def _init_communes():
    codes = set(SuiviBaldiya.objects.values_list('commune', flat=True))
    for code, _ in COMMUNE_CHOICES:
        if code not in codes:
            SuiviBaldiya.objects.create(commune=code)


@login_required
def tableau_bord(request):
    _init_communes()

    communes = SuiviBaldiya.objects.select_related('attribue_a', 'modifie_par').order_by('commune')

    filtre_statut = request.GET.get('statut', '')
    filtre_user   = request.GET.get('user', '')
    filtre_search = request.GET.get('q', '')

    if filtre_statut:
        communes = communes.filter(statut=filtre_statut)
    if filtre_user == 'moi':
        communes = communes.filter(attribue_a=request.user)
    elif filtre_user == 'non_assigne':
        communes = communes.filter(attribue_a__isnull=True)
    if filtre_search:
        matching = [c for c, n in COMMUNE_CHOICES if filtre_search in n]
        communes = communes.filter(commune__in=matching)

    all_c = SuiviBaldiya.objects.all()
    stats = {s[0]: all_c.filter(statut=s[0]).count() for s in STATUT_CHOICES}
    stats['total'] = all_c.count()

    from django.contrib.auth.models import User
    utilisateurs = User.objects.filter(is_active=True).order_by('username')

    # إضافة إحصائيات المهام لكل بلدية
    communes_with_tasks = []
    for commune in communes:
        # الحصول على إحصائيات المهام لهذه البلدية
        task_stats = SuiviTache.objects.filter(baldiya=commune).aggregate(
            total=models.Count('id'),
            termine=models.Count('id', filter=models.Q(statut='termine')),
            en_cours=models.Count('id', filter=models.Q(statut='en_cours')),
            probleme=models.Count('id', filter=models.Q(statut='probleme')),
        )
        task_stats['en_attente'] = task_stats['total'] - task_stats['termine'] - task_stats['en_cours'] - task_stats['probleme']

        communes_with_tasks.append({
            'commune': commune,
            'task_stats': task_stats,
        })

    statuts_json = json.dumps(list(STATUT_CHOICES))

    return render(request, 'suivi/tableau_bord.html', {
        'communes': communes_with_tasks,
        'stats': stats,
        'statuts': STATUT_CHOICES,
        'statuts_json': statuts_json,
        'utilisateurs': utilisateurs,
        'filtre_statut': filtre_statut,
        'filtre_user': filtre_user,
        'filtre_search': filtre_search,
    })


@login_required
def modifier_commune(request, pk):
    commune = get_object_or_404(SuiviBaldiya, pk=pk)
    if (commune.attribue_a and
            commune.attribue_a != request.user and
            not request.user.is_superuser):
        messages.warning(request,
            f'هذه البلدية مسندة بالفعل إلى {commune.attribue_a.username}.')
        return redirect('suivi:tableau_bord')

    if request.method == 'POST':
        form = SuiviBaldiyaForm(request.POST, instance=commune)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.modifie_par = request.user
            obj.save()
            messages.success(request, f'تم تحديث بلدية {commune.get_commune_display()} بنجاح ✓')
            return redirect('suivi:tableau_bord')
    else:
        form = SuiviBaldiyaForm(instance=commune)

    return render(request, 'suivi/modifier.html', {'form': form, 'commune': commune})


@login_required
@require_POST
def mise_a_jour_rapide(request, pk):
    commune = get_object_or_404(SuiviBaldiya, pk=pk)
    if (commune.attribue_a and
            commune.attribue_a != request.user and
            not request.user.is_superuser):
        return JsonResponse({'ok': False, 'msg': 'غير مصرح لك بتعديل هذه البلدية'}, status=403)

    data  = json.loads(request.body)
    champ = data.get('champ')
    valeur = data.get('valeur', '')

    if champ not in ['statut', 'remarque', 'date_contact']:
        return JsonResponse({'ok': False, 'msg': 'حقل غير صالح'}, status=400)

    setattr(commune, champ, valeur if valeur else None)
    commune.modifie_par = request.user
    commune.save()

    return JsonResponse({
        'ok':           True,
        'statut':       commune.statut,
        'statut_label': commune.get_statut_display(),
        'statut_color': commune.statut_color(),
        'statut_icon':  commune.statut_icon(),
        'modifie_par':  request.user.username,
        'modifie_le':   commune.modifie_le.strftime('%d/%m/%Y %H:%M'),
    })


@login_required
@require_POST
def assigner_commune(request, pk):
    commune = get_object_or_404(SuiviBaldiya, pk=pk)
    if commune.attribue_a and not request.user.is_superuser:
        return JsonResponse({
            'ok': False,
            'msg': f'هذه البلدية مسندة بالفعل إلى {commune.attribue_a.username}'
        }, status=403)

    data   = json.loads(request.body)
    action = data.get('action', 'claim')

    if action == 'claim':
        commune.attribue_a = request.user
    elif action == 'release' and (commune.attribue_a == request.user or request.user.is_superuser):
        commune.attribue_a = None

    commune.modifie_par = request.user
    commune.save()

    return JsonResponse({
        'ok':         True,
        'attribue_a': commune.attribue_a.username if commune.attribue_a else None,
    })


@login_required
def exporter_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, 'مكتبة openpyxl غير مثبتة. قم بتشغيل: pip install openpyxl')
        return redirect('suivi:tableau_bord')

    communes = SuiviBaldiya.objects.select_related('attribue_a', 'modifie_par').order_by('commune')

    # Appliquer les mêmes filtres que le tableau
    filtre_statut = request.GET.get('statut', '')
    filtre_user   = request.GET.get('user', '')
    if filtre_statut:
        communes = communes.filter(statut=filtre_statut)
    if filtre_user == 'moi':
        communes = communes.filter(attribue_a=request.user)
    elif filtre_user == 'non_assigne':
        communes = communes.filter(attribue_a__isnull=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "متابعة بلديات بجاية"
    ws.sheet_view.rightToLeft = True

    # Styles
    thin        = Side(border_style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Status colors (fill)
    statut_fills = {
        'en_attente': PatternFill("solid", fgColor="E9ECEF"),
        'contacte':   PatternFill("solid", fgColor="CFF4FC"),
        'en_cours':   PatternFill("solid", fgColor="FFF3CD"),
        'termine':    PatternFill("solid", fgColor="D1E7DD"),
        'probleme':   PatternFill("solid", fgColor="F8D7DA"),
    }
    statut_fonts = {
        'en_attente': Font(color="6C757D", size=10),
        'contacte':   Font(color="055160", size=10),
        'en_cours':   Font(color="664D03", size=10),
        'termine':    Font(color="0A3622", size=10),
        'probleme':   Font(color="58151C", size=10),
    }

    # ── Header row ────────────────────────────────────────────────────
    headers    = ['#', 'رمز البلدية', 'اسم البلدية', 'الحالة',
                  'مسند إلى', 'تاريخ الاتصال', 'الملاحظة',
                  'آخر تعديل بواسطة', 'تاريخ آخر تعديل']
    col_widths = [5, 12, 20, 18, 16, 16, 40, 18, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────
    statut_labels = dict(STATUT_CHOICES)

    for row_num, c in enumerate(communes, 2):
        sfill = statut_fills.get(c.statut, PatternFill("solid", fgColor="FFFFFF"))
        sfont = statut_fonts.get(c.statut, Font(size=10))

        row_data = [
            row_num - 1,
            c.commune,
            c.get_commune_display(),
            statut_labels.get(c.statut, c.statut),
            c.attribue_a.username if c.attribue_a else '—',
            c.date_contact.strftime('%d/%m/%Y') if c.date_contact else '—',
            c.remarque or '—',
            c.modifie_par.username if c.modifie_par else '—',
            c.modifie_le.strftime('%d/%m/%Y %H:%M') if c.modifie_le else '—',
        ]

        for col_num, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_num, column=col_num, value=value)
            cell.border    = border
            cell.alignment = right_align if col_num > 2 else center

            # Colorier la colonne "الحالة" selon le statut
            if col_num == 4:
                cell.fill = sfill
                cell.font = sfont
            else:
                alt = "F8FAFC" if row_num % 2 == 0 else "FFFFFF"
                cell.fill = PatternFill("solid", fgColor=alt)
                cell.font = Font(size=10)

        ws.row_dimensions[row_num].height = 20

    # ── Summary sheet ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("ملخص الإحصائيات")
    ws2.sheet_view.rightToLeft = True

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 16

    ws2.cell(row=1, column=1, value="الحالة").font       = header_font
    ws2.cell(row=1, column=1).fill                        = header_fill
    ws2.cell(row=1, column=1).alignment                   = center
    ws2.cell(row=1, column=2, value="العدد").font         = header_font
    ws2.cell(row=1, column=2).fill                        = header_fill
    ws2.cell(row=1, column=2).alignment                   = center
    ws2.cell(row=1, column=3, value="النسبة %").font      = header_font
    ws2.cell(row=1, column=3).fill                        = header_fill
    ws2.cell(row=1, column=3).alignment                   = center

    total = SuiviBaldiya.objects.count()
    for i, (code, label) in enumerate(STATUT_CHOICES, 2):
        count = SuiviBaldiya.objects.filter(statut=code).count()
        pct   = round(count / total * 100, 1) if total else 0
        ws2.cell(row=i, column=1, value=label).alignment  = right_align
        ws2.cell(row=i, column=2, value=count).alignment  = center
        ws2.cell(row=i, column=3, value=f"{pct}%").alignment = center
        ws2.cell(row=i, column=1).fill = statut_fills.get(code, PatternFill())
        for col in range(1, 4):
            ws2.cell(row=i, column=col).border = border

    ws2.cell(row=7, column=1, value="الإجمالي").font      = Font(bold=True, size=11)
    ws2.cell(row=7, column=2, value=total).font            = Font(bold=True, size=11)
    ws2.cell(row=7, column=3, value="100%").font           = Font(bold=True, size=11)
    for col in range(1, 4):
        ws2.cell(row=7, column=col).alignment = center
        ws2.cell(row=7, column=col).border    = border

    ws.freeze_panes  = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="متابعة_بلديات_بجاية.xlsx"'
    return response


@login_required
def gestion_taches(request):
    """إدارة المهام - عرض وإضافة المهام"""
    taches = Tache.objects.filter(actif=True).order_by('ordre')

    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        description = request.POST.get('description', '').strip()

        if nom:
            ordre_max = Tache.objects.aggregate(max_ordre=models.Max('ordre'))['max_ordre'] or 0
            Tache.objects.create(
                nom=nom,
                description=description,
                ordre=ordre_max + 1
            )
            messages.success(request, f'تم إضافة المهمة "{nom}" بنجاح ✓')
            return redirect('suivi:gestion_taches')
        else:
            messages.error(request, 'يرجى إدخال اسم المهمة')

    return render(request, 'suivi/gestion_taches.html', {
        'taches': taches,
    })


@login_required
@require_POST
def supprimer_tache(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'ok': False, 'msg': 'غير مصرح لك بهذا الإجراء'}, status=403)

    tache = get_object_or_404(Tache, pk=pk)
    nom = tache.nom
    tache.delete()

    messages.success(request, f'تم حذف المهمة "{nom}" بنجاح')
    return JsonResponse({'ok': True})


@login_required
def suivi_taches_commune(request, commune_pk):
    """متابعة المهام للبلدية المحددة"""
    commune = get_object_or_404(SuiviBaldiya, pk=commune_pk)
    taches = Tache.objects.filter(actif=True).order_by('ordre')

    # الحصول على متابعة المهام الموجودة أو إنشاؤها
    suivi_taches = []
    for tache in taches:
        suivi, created = SuiviTache.objects.get_or_create(
            baldiya=commune,
            tache=tache,
            defaults={'statut': 'en_attente'}
        )
        suivi_taches.append(suivi)

    if request.method == 'POST':
        tache_id = request.POST.get('tache_id')
        statut = request.POST.get('statut')
        remarque = request.POST.get('remarque', '').strip()

        if tache_id and statut:
            suivi = get_object_or_404(SuiviTache, baldiya=commune, tache_id=tache_id)

            # التحقق من الصلاحيات
            if (suivi.attribue_a and suivi.attribue_a != request.user and not request.user.is_superuser):
                messages.error(request, 'غير مصرح لك بتعديل هذه المهمة')
                return redirect('suivi:suivi_taches_commune', commune_pk=commune.pk)

            suivi.statut = statut
            suivi.remarque = remarque
            suivi.modifie_par = request.user

            # تحديث تاريخ البداية والانتهاء
            if statut in ['en_cours', 'termine'] and not suivi.date_debut:
                suivi.date_debut = timezone.now().date()
            if statut == 'termine' and not suivi.date_fin:
                suivi.date_fin = timezone.now().date()

            suivi.save()
            messages.success(request, f'تم تحديث مهمة "{suivi.tache.nom}" بنجاح ✓')

        return redirect('suivi:suivi_taches_commune', commune_pk=commune.pk)

    return render(request, 'suivi/suivi_taches_commune.html', {
        'commune': commune,
        'suivi_taches': suivi_taches,
        'statuts': STATUT_CHOICES,
    })


@login_required
@require_POST
def assigner_tache(request, commune_pk, tache_pk):
    commune = get_object_or_404(SuiviBaldiya, pk=commune_pk)
    tache = get_object_or_404(Tache, pk=tache_pk)

    suivi, created = SuiviTache.objects.get_or_create(
        baldiya=commune,
        tache=tache,
        defaults={'statut': 'en_attente'}
    )

    data = json.loads(request.body)
    action = data.get('action', 'claim')

    if action == 'claim':
        if suivi.attribue_a and not request.user.is_superuser:
            return JsonResponse({
                'ok': False,
                'msg': f'هذه المهمة مسندة بالفعل إلى {suivi.attribue_a.username}'
            }, status=403)
        suivi.attribue_a = request.user
    elif action == 'release' and (suivi.attribue_a == request.user or request.user.is_superuser):
        suivi.attribue_a = None

    suivi.modifie_par = request.user
    suivi.save()

    return JsonResponse({
        'ok': True,
        'attribue_a': suivi.attribue_a.username if suivi.attribue_a else None,
    })


@login_required
def rapport_taches(request):
    """تقرير شامل عن متابعة المهام"""
    communes = SuiviBaldiya.objects.select_related('attribue_a').order_by('commune')
    taches = Tache.objects.filter(actif=True).order_by('ordre')

    # إحصائيات المهام
    stats_taches = {}
    for tache in taches:
        suivi_tache = SuiviTache.objects.filter(tache=tache)
        stats_taches[tache.id] = {
            'tache': tache,
            'total': suivi_tache.count(),
            'en_attente': suivi_tache.filter(statut='en_attente').count(),
            'en_cours': suivi_tache.filter(statut='en_cours').count(),
            'termine': suivi_tache.filter(statut='termine').count(),
            'probleme': suivi_tache.filter(statut='probleme').count(),
        }

    # بيانات المتابعة لكل بلدية
    suivi_data = {}
    for commune in communes:
        # الحصول على جميع متابعات المهام لهذه البلدية
        suivi_objects = SuiviTache.objects.filter(baldiya=commune).select_related('tache', 'attribue_a')

        # إنشاء قائمة من المهام مع متابعاتها
        taches_suivi = []
        for tache in taches:
            suivi = next((s for s in suivi_objects if s.tache_id == tache.id), None)
            taches_suivi.append({
                'tache': tache,
                'suivi': suivi,
            })

        suivi_data[commune.pk] = {
            'commune': commune,
            'taches_suivi': taches_suivi,
        }

    return render(request, 'suivi/rapport_taches.html', {
        'taches': taches,
        'suivi_data': suivi_data,
        'stats_taches': stats_taches,
        'statuts': STATUT_CHOICES,
    })


@login_required
def get_commune_tasks_api(request, commune_pk):
    """API للحصول على مهام البلدية بتنسيق JSON"""
    commune = get_object_or_404(SuiviBaldiya, pk=commune_pk)
    taches = Tache.objects.filter(actif=True).order_by('ordre')

    tasks_data = []
    for tache in taches:
        suivi, created = SuiviTache.objects.get_or_create(
            baldiya=commune,
            tache=tache,
            defaults={'statut': 'en_attente'}
        )

        statut_colors = {
            'en_attente': 'secondary',
            'en_cours': 'warning',
            'termine': 'success',
            'probleme': 'danger'
        }

        tasks_data.append({
            'id': suivi.id,
            'titre': tache.nom,
            'description': tache.description,
            'statut': suivi.get_statut_display(),
            'statut_code': suivi.statut,
            'statut_color': statut_colors.get(suivi.statut, 'secondary'),
            'assigned': suivi.attribue_a.username if suivi.attribue_a else None,
            'date_debut': suivi.date_debut.strftime('%d/%m/%Y') if suivi.date_debut else '-',
            'date_fin': suivi.date_fin.strftime('%d/%m/%Y') if suivi.date_fin else '-',
            'remarque': suivi.remarque or '',
        })

    return JsonResponse({'tasks': tasks_data})


@login_required
@require_POST
def update_task_status(request, commune_pk, task_pk):
    """API لتحديث حالة المهمة"""
    commune = get_object_or_404(SuiviBaldiya, pk=commune_pk)
    suivi = get_object_or_404(SuiviTache, baldiya=commune, pk=task_pk)

    # التحقق من الصلاحيات
    if (suivi.attribue_a and suivi.attribue_a != request.user and not request.user.is_superuser):
        return JsonResponse({'ok': False, 'msg': 'غير مصرح لك بتعديل هذه المهمة'}, status=403)

    data = json.loads(request.body)
    statut = data.get('statut')
    remarque = data.get('remarque', '').strip()

    if statut not in dict(STATUT_CHOICES):
        return JsonResponse({'ok': False, 'msg': 'حالة غير صالحة'}, status=400)

    suivi.statut = statut
    suivi.remarque = remarque
    suivi.modifie_par = request.user

    # تحديث التواريخ
    if statut in ['en_cours', 'termine'] and not suivi.date_debut:
        suivi.date_debut = timezone.now().date()
    if statut == 'termine' and not suivi.date_fin:
        suivi.date_fin = timezone.now().date()

    suivi.save()

    statut_colors = {
        'en_attente': 'secondary',
        'en_cours': 'warning',
        'termine': 'success',
        'probleme': 'danger'
    }

    return JsonResponse({
        'ok': True,
        'statut': suivi.get_statut_display(),
        'statut_color': statut_colors.get(suivi.statut, 'secondary'),
        'date_debut': suivi.date_debut.strftime('%d/%m/%Y') if suivi.date_debut else '-',
        'date_fin': suivi.date_fin.strftime('%d/%m/%Y') if suivi.date_fin else '-',
    })


@login_required
@require_POST
def delete_task(request, commune_pk, task_pk):
    """API لحذف متابعة المهمة"""
    if not request.user.is_superuser:
        return JsonResponse({'ok': False, 'msg': 'غير مصرح لك بهذا الإجراء'}, status=403)

    commune = get_object_or_404(SuiviBaldiya, pk=commune_pk)
    suivi = get_object_or_404(SuiviTache, baldiya=commune, pk=task_pk)

    suivi.delete()

    return JsonResponse({'ok': True})
